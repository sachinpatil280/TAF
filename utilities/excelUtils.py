import openpyxl


class ReadExcel:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = self.get_workbook()
        self.sheet_name = None

    def get_workbook(self):
        return openpyxl.load_workbook(self.excel_file)

    def get_row_count(self, sheet_name):
        self.sheet_name = sheet_name
        sheet = self.get_sheet()
        return sheet.max_row

    def get_column_count(self, sheet_name):
        self.sheet_name = sheet_name
        sheet = self.get_sheet()
        return sheet.max_column

    def read_data(self, sheet_name, row_num, col_num):
        self.sheet_name = sheet_name
        sheet = self.get_sheet()
        return sheet.cell(row=row_num, column=col_num).value

    def write_data(self, sheet_name, row_num, col_num, data):
        self.sheet_name = sheet_name
        sheet = self.get_sheet()
        sheet.cell(row=row_num, column=col_num).value = data
        self.workbook.save(self.excel_file)

    @property
    def get_sheet(self):
        if self.sheet_name:
            return self.workbook[self.sheet_name]
        else:
            raise TypeError
